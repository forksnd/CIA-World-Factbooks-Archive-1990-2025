"""Generate raw-sources/INDEX.xlsx (3 sheets: Files, Summary, How to verify)
from MANIFEST.json. The xlsx ships inside the Release bundle as the
human-browsable view; the bundle's README.md is the human-readable summary.
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
BUNDLE = REPO / "raw-sources"
MANIFEST = BUNDLE / "MANIFEST.json"
XLSX = BUNDLE / "INDEX.xlsx"


def build_xlsx(manifest):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Files ----
    ws = wb.active
    ws.title = "Files"
    headers = ["Year", "Era", "Filename", "Size (MB)", "SHA256", "Upstream", "Parser script", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F343C")
        c.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, e in enumerate(manifest["files"], 2):
        # Upstream URL: html/text use url, json uses repo+commit
        if "upstream_url" in e:
            upstream = e["upstream_url"]
        else:
            upstream = f"{e.get('upstream_repo', '')} @ {e.get('upstream_commit_short', '')} ({e.get('upstream_commit_date', '')})"
        ws.cell(row=row_idx, column=1, value=e["year"])
        ws.cell(row=row_idx, column=2, value=e["era"])
        ws.cell(row=row_idx, column=3, value=e["filename"])
        ws.cell(row=row_idx, column=4, value=round(e["size_bytes"] / 1e6, 2))
        ws.cell(row=row_idx, column=5, value=e["sha256"])
        link_cell = ws.cell(row=row_idx, column=6, value=upstream)
        if upstream.startswith("http"):
            link_cell.hyperlink = upstream
            link_cell.font = Font(color="3B82F6", underline="single")
        ws.cell(row=row_idx, column=7, value=e["parser_script"])
        ws.cell(row=row_idx, column=8, value=e.get("notes", ""))

    widths = [7, 16, 28, 12, 70, 80, 38, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Sheet 2: Summary ----
    ws2 = wb.create_sheet("Summary")
    rows = [
        ["Era", "Year range", "Files", "Total size (MB)", "Source"],
        ["Gutenberg plaintext", "1990-1999, 2001 + 1996_cia_original", "12", "38",
         "Project Gutenberg + Wayback ODCI 1997-05-28"],
        ["Wayback HTML zips", "2000-2020", "21", "2,810",
         "Wayback Machine captures of cia.gov factbook archive page"],
        ["JSON snapshots", "2021-2025", "5", "29",
         "github.com/factbook/cache.factbook.json (year-end commits)"],
        [],
        ["Total", "1990-2025 (36 years)", "38", "2,978", ""],
    ]
    for row_idx, r in enumerate(rows, 1):
        for col_idx, v in enumerate(r, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=v)
            if row_idx == 1 or (row_idx == 6 and v):
                c.font = Font(bold=True, color="FFFFFF" if row_idx == 1 else "000000")
                if row_idx == 1:
                    c.fill = PatternFill("solid", fgColor="2F343C")
    for i, w in enumerate([28, 38, 8, 18, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 3: How to verify ----
    ws3 = wb.create_sheet("How to verify")
    notes = [
        ["CIA World Factbook Archive 1990-2025 — Raw Sources"],
        [""],
        ["Every file in this bundle has a SHA256 hash in MANIFEST.json."],
        [""],
        ["To verify all files at once:"],
        [""],
        ["Linux / macOS:"],
        ["  cd raw-sources"],
        ["  python -c \"import json,hashlib,os; m=json.load(open('MANIFEST.json'));"],
        ["    [print('OK' if hashlib.sha256(open(next(p for p in ["],
        ["      f'html/{e[\\\"filename\\\"]}',f'text/{e[\\\"filename\\\"]}',f'json/{e[\\\"filename\\\"]}'"],
        ["    ] if os.path.exists(p)),'rb').read()).hexdigest()==e['sha256'] else 'FAIL',"],
        ["    e['filename']) for e in m['files']]\""],
        [""],
        ["Windows PowerShell (per file):"],
        ["  Get-FileHash html\\factbook-2010.zip -Algorithm SHA256"],
        [""],
        ["Provenance: 99.94% of records in factbook.db CountryFields trace back"],
        ["to these raw files via a row-level diff. See raw-sources/VALIDATION.md"],
        ["in the main repo for the full validation methodology."],
    ]
    for row_idx, r in enumerate(notes, 1):
        if r:
            ws3.cell(row=row_idx, column=1, value=r[0])
            if row_idx == 1:
                ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    ws3.column_dimensions["A"].width = 100

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"  wrote {XLSX}")


def main():
    with open(MANIFEST, "r", encoding="utf-8") as f:
        manifest = json.load(f)
    print("Building manifest views...")
    build_xlsx(manifest)
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
